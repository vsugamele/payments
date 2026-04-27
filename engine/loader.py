"""
Carrega regras e taxas dos arquivos Excel para uso na engine de provisionamento.
"""
import re
import warnings
import openpyxl
from pathlib import Path
from typing import NamedTuple, Optional
import datetime as _dt

BASE_DIR = Path(__file__).parent.parent
RULES_FILE = BASE_DIR / "Novo_Intercambio.xlsx"
RATES_FILE = BASE_DIR / "Tabelas" / "tabela_Intercambio Master.xlsx"
VISA_RULES_FILE = BASE_DIR / "Novo_visa_14_v2.xlsx"


# ---------------------------------------------------------------------------
# Estruturas de dados
# ---------------------------------------------------------------------------

class MCRule(NamedTuple):
    rule_id: int
    expression: str
    ird: str
    priority: int
    active: bool
    description: str
    compiled: object  # código compilado para eval


class MaestroRule(NamedTuple):
    rule_id: int
    expression: str
    pseudo_ird: str
    priority: int
    rate_pct: float
    cap_brl: Optional[float]  # teto em R$ (ex: R$0,30 para MSI_CAP)
    description: str
    compiled: object


# ---------------------------------------------------------------------------
# Compilador de expressões
# Converte sintaxe MC: ((PRODUCT_TYPE="CREDIT")&(CARD_PRESENT_ID=0))
# Para Python eval: (ctx.get('PRODUCT_TYPE') == 'CREDIT') and (ctx.get('CARD_PRESENT_ID') == 0)
# ---------------------------------------------------------------------------

def _compile_expr(expr: str) -> object:
    py = expr.replace('&', ' and ').replace('|', ' or ')

    # (VAR="STRING") → (ctx.get('VAR') == 'STRING')
    py = re.sub(r'\((\w+)="([^"]+)"\)', r"(ctx.get('\1') == '\2')", py)

    # (VAR>=NUMBER)
    py = re.sub(r'\((\w+)>=([\d.]+)\)', r"(ctx.get('\1', 0) >= \2)", py)

    # (VAR<=NUMBER)
    py = re.sub(r'\((\w+)<=([\d.]+)\)', r"(ctx.get('\1', 0) <= \2)", py)

    # (VAR=NUMBER)
    py = re.sub(r'\((\w+)=(\d+)\)', r"(ctx.get('\1') == \2)", py)

    return compile(py, "<expr>", "eval")


def _eval_expr(compiled, ctx: dict) -> bool:
    try:
        return bool(eval(compiled, {"__builtins__": {}}, {"ctx": ctx}))
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Loaders
# ---------------------------------------------------------------------------

def load_mc_rules() -> list[MCRule]:
    wb = openpyxl.load_workbook(RULES_FILE, data_only=True)
    ws = wb["Mastercard"]
    rules: list[MCRule] = []
    skip_header = True

    for row in ws.iter_rows(values_only=True):
        if skip_header:
            skip_header = False
            continue
        if not row[0] or not row[2] or not row[3]:
            continue

        rule_id, _, expression, ird, priority, active = row[0], row[1], row[2], row[3], row[4], row[5]
        description = row[8] if len(row) > 8 else ""

        active_bool = bool(active) if active is not None else True
        if not active_bool:
            continue

        try:
            compiled = _compile_expr(str(expression))
        except Exception:
            continue

        rules.append(MCRule(
            rule_id=int(rule_id),
            expression=str(expression),
            ird=str(ird).strip(),
            priority=int(priority) if priority else 999,
            active=active_bool,
            description=str(description) if description else "",
            compiled=compiled,
        ))

    rules.sort(key=lambda r: r.priority)
    return rules


def load_maestro_rules() -> list[MaestroRule]:
    wb = openpyxl.load_workbook(RULES_FILE, data_only=True)
    ws = wb["Maestro"]
    rules: list[MaestroRule] = []
    skip_header = True

    for row in ws.iter_rows(values_only=True):
        if skip_header:
            skip_header = False
            continue
        if not row[0] or not row[1]:
            continue

        rule_id, expression, pseudo_ird, priority, taxa, desc = (
            row[0], row[1], row[2], row[3], row[4], row[5],
        )

        rate_str = str(taxa).replace("%", "").strip() if taxa else "0"
        try:
            rate_pct = float(rate_str)
        except ValueError:
            rate_pct = 0.0

        cap = 0.30 if pseudo_ird == "MSI_CAP" else None

        try:
            compiled = _compile_expr(str(expression))
        except Exception:
            continue

        rules.append(MaestroRule(
            rule_id=int(rule_id),
            expression=str(expression),
            pseudo_ird=str(pseudo_ird),
            priority=int(priority) if priority else 999,
            rate_pct=rate_pct,
            cap_brl=cap,
            description=str(desc) if desc else "",
            compiled=compiled,
        ))

    rules.sort(key=lambda r: r.priority)
    return rules


# ---------------------------------------------------------------------------
# Tabela de taxas
#
# Estrutura Debito e Credito:
#   col0: Tabela (Table 76 = Consumer credit, Table 79 = Commercial, etc.)
#   col1: Cobrança
#   col2: Descrição ("Basic rate" ou "Adjustments to the Basic Rate")
#   col3: transaction_type
#   col4: IRD
#   col5: tier
#   col6: segment
#   col7: taxa_base (% direta, para "Basic rate")
#   col8: acrescimo (ajuste sobre IA, para "Adjustments")
#   col9: desconto
#
# Lógica: IRDs com "Basic rate" têm taxa direta.
#         IRDs com "Adjustments" usam: taxa_IA(tier, segment) + acrescimo
#
# Chave primária de lookup: (ird, tier, segment)
# ---------------------------------------------------------------------------

def load_rates() -> tuple[dict, dict]:
    """
    Retorna:
        base_rates: {(ird, tier, segment): rate_pct}  — taxas diretas (IRD IA, JA, IV, IW, etc.)
        adjustments: {(ird, segment): adjustment_pct} — ajustes sobre IA para IRDs CNP
    """
    wb = openpyxl.load_workbook(RATES_FILE, data_only=True)
    ws = wb["Debito e Credito"]

    base_rates: dict[tuple, float] = {}
    adjustments: dict[tuple, float] = {}

    skip_header = True
    for row in ws.iter_rows(values_only=True):
        if skip_header:
            skip_header = False
            continue
        if not any(row):
            continue

        descricao = str(row[2]) if row[2] else ""
        ird = str(row[4]).strip() if row[4] else None
        tier = str(row[5]).strip() if row[5] else None
        segment = str(row[6]).strip() if row[6] else None

        if not ird or not tier or not segment:
            continue

        if "Basic rate" in descricao:
            taxa = row[7]
            if taxa is not None and isinstance(taxa, (int, float)):
                base_rates[(ird, tier, segment)] = float(taxa)

        elif "Adjustments" in descricao:
            acrescimo = row[8]
            desconto = row[9]
            if acrescimo is not None and isinstance(acrescimo, (int, float)):
                adjustments[(ird, segment)] = float(acrescimo)
            elif desconto is not None and isinstance(desconto, (int, float)):
                adjustments[(ird, segment)] = -float(desconto)

    return base_rates, adjustments


def load_mcc_to_segment() -> dict[int, str]:
    """Mapeia MCC → segmento Mastercard (SprMkt, GenEx, CommT, etc.)"""
    wb = openpyxl.load_workbook(RATES_FILE, data_only=True)
    ws = wb["MCC"]

    # Mapeia categoria textual → código de segmento
    CAT_TO_SEG = {
        "supermarkets": "SprMkt",
        "general expenses": "GenEx",
        "government services": "GovSvc",
        "wholesalers": "Whole",
        "micro merchant": "MM",
        "commuter transport": "CommT",
        "t&e": "T&E",
        "lottery": "Lott",
    }

    mcc_seg: dict[int, str] = {}
    skip_header = True
    for row in ws.iter_rows(values_only=True):
        if skip_header:
            skip_header = False
            continue
        if not row[0] or not row[1]:
            continue
        mcc = row[0]
        cat = str(row[1]).strip().lower()
        seg = CAT_TO_SEG.get(cat, "Other")
        try:
            mcc_seg[int(mcc)] = seg
        except (TypeError, ValueError):
            pass

    return mcc_seg


def eval_rule(rule, ctx: dict) -> bool:
    return _eval_expr(rule.compiled, ctx)


# ---------------------------------------------------------------------------
# Visa
# ---------------------------------------------------------------------------

class VisaRule(NamedTuple):
    rule_id: str        # ex: 'BR-ECOM-PLT-SEC'
    priority: int
    descriptor: str     # descrição humana
    accounting_sign: str  # 'PAY_ISS' = despesa, 'RCV_ISS' = receita
    rate_pct: float
    fixed_fee: float
    cap_fee: float
    expression: str
    compiled: object
    rate_corrupted: bool = False  # True quando a célula Excel tinha uma data em vez de número


def _compile_visa_expr(expr: str) -> object:
    """
    Converte sintaxe Visa → Python eval.

    Padrões suportados:
      (VAR IN ('a','b'))     → (ctx.get('VAR') in ('a','b'))
      (VAR='str')            → (ctx.get('VAR') == 'str')
      (VAR!='str')           → (ctx.get('VAR') != 'str')
      (VAR >= N)             → (ctx.get('VAR', 0) >= N)
      (VAR <= N)             → (ctx.get('VAR', 0) <= N)
      (VAR > N)              → (ctx.get('VAR', 0) > N)
      (VAR < N)              → (ctx.get('VAR', 0) < N)
      (VAR=N)                → (ctx.get('VAR') == N)   -- número inteiro
      (VAR IS NOT NULL)      → (ctx.get('VAR') not in (None, ''))
      (VAR IS EMPTY)         → (ctx.get('VAR') in (None, ''))
      &                      → and
      |                      → or
    """
    py = expr

    # IS NOT NULL / IS EMPTY (antes de qualquer outro padrão)
    py = re.sub(r'\((\w+) IS NOT NULL\)', r"(ctx.get('\1') not in (None, ''))", py)
    py = re.sub(r'\((\w+) IS EMPTY\)', r"(ctx.get('\1') in (None, ''))", py)

    # IN ('a','b','c')
    def _replace_in(m: re.Match) -> str:
        var = m.group(1)
        items = m.group(2)
        return f"(ctx.get('{var}') in ({items}))"

    py = re.sub(r"\((\w+) IN \(([^)]+)\)\)", _replace_in, py)

    # Comparações numéricas (>=, <=, >, <)
    py = re.sub(r'\((\w+) >= ([\d.]+)\)', r"(ctx.get('\1', 0) >= \2)", py)
    py = re.sub(r'\((\w+) <= ([\d.]+)\)', r"(ctx.get('\1', 0) <= \2)", py)
    py = re.sub(r'\((\w+) > ([\d.]+)\)', r"(ctx.get('\1', 0) > \2)", py)
    py = re.sub(r'\((\w+) < ([\d.]+)\)', r"(ctx.get('\1', 0) < \2)", py)

    # Igualdade com string: (VAR='str')
    py = re.sub(r"\((\w+)='([^']+)'\)", r"(ctx.get('\1') == '\2')", py)

    # Desigualdade com string: (VAR!='str')
    py = re.sub(r"\((\w+)!='([^']+)'\)", r"(ctx.get('\1') != '\2')", py)

    # Igualdade com número inteiro: (VAR=N)
    py = re.sub(r'\((\w+)=(\d+)\)', r"(ctx.get('\1') == \2)", py)

    # Operadores booleanos
    py = py.replace('&', ' and ').replace('|', ' or ')

    return compile(py, "<visa_expr>", "eval")


def load_visa_rules() -> list[VisaRule]:
    """Carrega e compila as regras Visa de Novo_visa_14_v2.xlsx."""
    wb = openpyxl.load_workbook(VISA_RULES_FILE, data_only=True)
    ws = wb.active
    rules: list[VisaRule] = []
    skip_header = True

    for row in ws.iter_rows(values_only=True):
        if skip_header:
            skip_header = False
            continue

        # Colunas: RULE_ID, PRIORITY, DESCRIPTOR, ACCOUNTING_SIGN,
        #          RATE_PCT, FIXED_FEE, CAP_FEE, LOGICAL_EXPRESSION
        rule_id = row[0]
        priority = row[1]
        descriptor = row[2]
        accounting_sign = row[3]
        rate_raw = row[4]
        fixed_fee_raw = row[5]
        cap_fee_raw = row[6]
        expression = row[7]

        if not rule_id or not expression:
            continue

        # Parse da taxa — célula pode estar formatada como data por engano no Excel
        rate_corrupted = False
        if isinstance(rate_raw, (_dt.datetime, _dt.date)):
            warnings.warn(
                f"[Visa loader] Regra '{rule_id}': RATE_PCT contém uma data ({rate_raw}) "
                "em vez de número. Célula provavelmente mal-formatada no Excel. "
                "Corrija o arquivo e recarregue. Taxa definida como 0.00% temporariamente.",
                UserWarning,
                stacklevel=2,
            )
            rate_pct = 0.0
            rate_corrupted = True
        else:
            try:
                rate_pct = float(str(rate_raw).replace("%", "").strip())
            except (ValueError, TypeError):
                rate_pct = 0.0  # regra sem taxa (ex: OCT/disputes a 0%)

        try:
            fixed_fee = float(str(fixed_fee_raw).replace("%", "").strip())
        except (ValueError, TypeError):
            fixed_fee = 0.0

        try:
            cap_fee = float(str(cap_fee_raw).replace("%", "").strip())
        except (ValueError, TypeError):
            cap_fee = 0.0

        try:
            compiled = _compile_visa_expr(str(expression))
        except Exception:
            continue  # expressão malformada — pula

        rules.append(VisaRule(
            rule_id=str(rule_id).strip(),
            priority=int(priority) if priority else 999,
            descriptor=str(descriptor).strip() if descriptor else "",
            accounting_sign=str(accounting_sign).strip() if accounting_sign else "PAY_ISS",
            rate_pct=rate_pct,
            fixed_fee=fixed_fee,
            cap_fee=cap_fee,
            expression=str(expression),
            compiled=compiled,
            rate_corrupted=rate_corrupted,
        ))

    rules.sort(key=lambda r: r.priority)
    return rules
