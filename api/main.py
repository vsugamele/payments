"""
API de Provisionamento de Intercâmbio - Mastercard / Maestro / Visa
"""
import io
import time
from pathlib import Path

import openpyxl
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse

from .schemas import (
    CalcTecnicoRequest,
    CalcSimplesRequest,
    CalcBatchRequest,
    CalcVisaTecnicoRequest,
    CalcVisaSimplesRequest,
)
from engine.calculator import InterchangeCalculator
from .routers import rag

app = FastAPI(
    title="Motor de Provisionamento de Intercâmbio",
    description=(
        "Calcula taxas de intercâmbio Mastercard, Maestro e Visa aplicando "
        "lógica de cascata (waterfall) por prioridade de regras."
    ),
    version="1.1.0",
)

app.include_router(rag.router, prefix="/api/rag", tags=["RAG"])


# CORS — permite que o frontend Next.js (localhost:3000) acesse a API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:3001", "http://192.168.0.3:3000", "http://192.168.0.3:3001"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Singleton — carrega as tabelas uma vez ao iniciar
_calculator: InterchangeCalculator | None = None
_loaded_at: float = 0.0


def get_calc() -> InterchangeCalculator:
    global _calculator, _loaded_at
    if _calculator is None:
        _calculator = InterchangeCalculator()
        _loaded_at = time.time()
    return _calculator


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.get("/")
def health():
    calc = get_calc()
    return {
        "status": "ok",
        "version": "1.1.0",
        "regras_mastercard": len(calc.mc_rules),
        "regras_maestro": len(calc.maestro_rules),
        "regras_visa": len(calc.visa_rules),
        "taxas_base": len(calc.base_rates),
        "loaded_at": _loaded_at,
    }


@app.post("/reload")
def reload_tables():
    """
    Recarrega todas as tabelas de regras do disco (Excel).
    Use após editar os arquivos .xlsx — o motor lê os dados atualizados
    sem precisar reiniciar o servidor.
    """
    global _calculator, _loaded_at
    _calculator = None
    calc = get_calc()  # força recarga imediata
    return {
        "status": "reloaded",
        "regras_mastercard": len(calc.mc_rules),
        "regras_maestro": len(calc.maestro_rules),
        "regras_visa": len(calc.visa_rules),
        "taxas_base": len(calc.base_rates),
        "loaded_at": _loaded_at,
    }


@app.post("/calcular/tecnico")
def calcular_tecnico(req: CalcTecnicoRequest):
    """
    Cálculo com parâmetros técnicos completos (ISO 8583 / IPM).
    Use quando já tiver os campos mapeados da mensagem de autorização.
    """
    params = req.model_dump()
    if not params.get("TXN_AMOUNT"):
        params["TXN_AMOUNT"] = params["valor"]
    result = get_calc().calcular(params)
    return result.to_dict()


@app.post("/calcular/simples")
def calcular_simples(req: CalcSimplesRequest):
    """
    Cálculo simplificado com termos em português.
    Ideal para simulações rápidas e testes de cenário.
    """
    result = get_calc().calcular_simples(req.model_dump())
    return result.to_dict()


@app.post("/calcular/lote")
def calcular_lote(req: CalcBatchRequest):
    """
    Cálculo em lote (JSON). Envia uma lista de transações e recebe
    todas as taxas calculadas de uma vez.
    """
    calc = get_calc()
    resultados = []
    for item in req.transacoes:
        try:
            if item.simples:
                r = calc.calcular_simples(item.params)
            else:
                params = item.params.copy()
                if not params.get("TXN_AMOUNT"):
                    params["TXN_AMOUNT"] = params.get("valor", 0)
                r = calc.calcular(params)
            d = r.to_dict()
        except Exception as e:
            d = {"sucesso": False, "erro": str(e)}

        if item.id:
            d["id"] = item.id
        resultados.append(d)

    return {"total": len(resultados), "resultados": resultados}


@app.post("/calcular/lote/excel")
async def calcular_lote_excel(arquivo: UploadFile = File(...)):
    """
    Upload de planilha Excel (.xlsx) com colunas de transações.
    Retorna uma planilha com a coluna `taxa_intercambio` preenchida.

    Colunas esperadas (modo simplificado):
      valor, bandeira, tipo_cartao, pessoa, canal, mcc, categoria, parcelas
    """
    contents = await arquivo.read()
    try:
        wb_in = openpyxl.load_workbook(io.BytesIO(contents))
    except Exception:
        raise HTTPException(status_code=400, detail="Arquivo Excel inválido.")

    ws = wb_in.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    calc = get_calc()

    # Adiciona colunas de resultado
    new_cols = ["ird", "descricao_regra", "segment", "rate_pct", "taxa_intercambio", "erro"]
    for col_name in new_cols:
        ws.cell(row=1, column=len(headers) + new_cols.index(col_name) + 1, value=col_name)

    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                row_data[header] = val

        if not row_data.get("valor"):
            continue

        try:
            # Normaliza tipos
            params = {
                "valor": float(row_data.get("valor", 0)),
                "bandeira": str(row_data.get("bandeira", "mastercard")).lower(),
                "tipo_cartao": str(row_data.get("tipo_cartao", "standard")).lower(),
                "pessoa": str(row_data.get("pessoa", "fisica")).lower(),
                "canal": str(row_data.get("canal", "fisico")).lower(),
                "parcelas": int(row_data.get("parcelas", 1)),
            }
            if row_data.get("mcc"):
                params["mcc"] = int(row_data["mcc"])
            if row_data.get("categoria"):
                params["categoria"] = str(row_data["categoria"])

            result = calc.calcular_simples(params)
            d = result.to_dict()
        except Exception as e:
            d = {"sucesso": False, "erro": str(e)}

        base_col = len(headers) + 1
        ws.cell(row=row_idx, column=base_col, value=d.get("ird", ""))
        ws.cell(row=row_idx, column=base_col + 1, value=d.get("descricao_regra", ""))
        ws.cell(row=row_idx, column=base_col + 2, value=d.get("segment", ""))
        ws.cell(row=row_idx, column=base_col + 3, value=d.get("rate_pct", ""))
        ws.cell(row=row_idx, column=base_col + 4, value=d.get("taxa_intercambio", ""))
        ws.cell(row=row_idx, column=base_col + 5, value=d.get("erro", ""))

    output = io.BytesIO()
    wb_in.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=resultado_intercambio.xlsx"},
    )


@app.post("/calcular/visa/tecnico")
def calcular_visa_tecnico(req: CalcVisaTecnicoRequest):
    """
    Cálculo Visa com parâmetros técnicos do Clearing.
    Ideal para homologação e testes com campos ISO/Visa reais.
    """
    result = get_calc().calcular_visa(req.model_dump())
    return result.to_dict()


@app.post("/calcular/visa/simples")
def calcular_visa_simples(req: CalcVisaSimplesRequest):
    """
    Cálculo Visa com termos simplificados em português.
    Ideal para simulações rápidas por produto/canal/categoria.
    """
    result = get_calc().calcular_visa_simples(req.model_dump())
    return result.to_dict()


@app.get("/regras/visa")
def listar_regras_visa():
    """Lista todas as regras Visa carregadas (ordenadas por prioridade)."""
    calc = get_calc()
    return {
        "total": len(calc.visa_rules),
        "regras": [
            {
                "rule_id": r.rule_id,
                "priority": r.priority,
                "descriptor": r.descriptor,
                "accounting_sign": r.accounting_sign,
                "rate_pct": r.rate_pct,
                "fixed_fee": r.fixed_fee,
                "cap_fee": r.cap_fee,
                "expression": r.expression,
                **({"aviso": "RATE_PCT corrompido na planilha (célula com data)"} if r.rate_corrupted else {}),
            }
            for r in calc.visa_rules
        ],
    }


@app.get("/regras/mastercard")
def listar_regras_mc(limit: int = 50, offset: int = 0):
    """Lista as regras Mastercard carregadas (ordenadas por prioridade)."""
    calc = get_calc()
    rules = calc.mc_rules[offset: offset + limit]
    return {
        "total": len(calc.mc_rules),
        "regras": [
            {"rule_id": r.rule_id, "ird": r.ird, "priority": r.priority, "description": r.description}
            for r in rules
        ],
    }


@app.get("/regras/maestro")
def listar_regras_maestro():
    """Lista as regras Maestro carregadas."""
    calc = get_calc()
    return {
        "total": len(calc.maestro_rules),
        "regras": [
            {
                "rule_id": r.rule_id,
                "pseudo_ird": r.pseudo_ird,
                "priority": r.priority,
                "rate_pct": r.rate_pct,
                "cap_brl": r.cap_brl,
                "description": r.description,
            }
            for r in calc.maestro_rules
        ],
    }


@app.get("/taxas/irds")
def listar_irds():
    """Lista todos os IRDs disponíveis na tabela de taxas."""
    calc = get_calc()
    irds = sorted({k[0] for k in calc.base_rates})
    return {"irds": irds}
