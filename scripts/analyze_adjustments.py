"""
Analise comparativa: ajustes do banco vs planilhas de origem.
"""
import sys
import os
import psycopg2
import openpyxl
from dotenv import load_dotenv

sys.stdout.reconfigure(encoding='utf-8')
load_dotenv()

conn = psycopg2.connect(os.getenv("DATABASE_URL"))
cursor = conn.cursor()

print("=" * 80)
print("ANALISE 1: Ajustes no banco (ic_adjustments)")
print("=" * 80)

# Busca todos os ajustes para segmento 'Base' para facilitar comparação
cursor.execute("""
    SELECT ird, segment, adjustment_pct
    FROM ic_adjustments
    WHERE segment IN ('Base', 'Other')
    ORDER BY ird, segment;
""")
db_adj = {(r[0], r[1]): r[2] for r in cursor.fetchall()}

print(f"\nAjustes em ic_adjustments (segment=Base/Other):")
for (ird, seg), adj in sorted(db_adj.items()):
    sinal = f"+{adj:.4f}%" if adj >= 0 else f"{adj:.4f}%"
    print(f"  {ird:6s}  seg={seg:8s}  ajuste={sinal}")

print("\n" + "=" * 80)
print("ANALISE 2: Planilha Mastercard_30_v2.xlsx — aba Variavel")
print("=" * 80)

wb = openpyxl.load_workbook("Mastercard_30_v2.xlsx", data_only=True)
ws_var = wb["Variavel"]

print("\nColunas detectadas na aba Variavel:")
headers = [ws_var.cell(1, c).value for c in range(1, 20) if ws_var.cell(1, c).value]
for i, h in enumerate(headers, 1):
    print(f"  Col {i}: {h}")

print("\nPrimeiras 30 linhas da aba Variavel:")
for row in ws_var.iter_rows(min_row=1, max_row=30, values_only=True):
    if any(v is not None for v in row):
        print(f"  {[str(v)[:20] if v is not None else '-' for v in row[:10]]}")

print("\n" + "=" * 80)
print("ANALISE 3: Planilha Novo_Intercambio.xlsx — aba Mastercard")
print("=" * 80)

wb2 = openpyxl.load_workbook("Novo_Intercambio.xlsx", data_only=True)
ws_mc = wb2["Mastercard"]

print("\nColunas detectadas:")
headers2 = [ws_mc.cell(1, c).value for c in range(1, 20) if ws_mc.cell(1, c).value]
for i, h in enumerate(headers2, 1):
    print(f"  Col {i}: {h}")

print("\nPrimeiras 40 linhas:")
for row in ws_mc.iter_rows(min_row=1, max_row=40, values_only=True):
    if any(v is not None for v in row):
        print(f"  {[str(v)[:25] if v is not None else '-' for v in row[:8]]}")

print("\n" + "=" * 80)
print("ANALISE 4: Taxas base x ajustes no banco — resumo normativo")
print("=" * 80)

# Busca taxa IA/Consumer standard/Base como referência
cursor.execute("""
    SELECT rate_pct FROM ic_base_rates
    WHERE ird='IA' AND tier='Consumer standard' AND segment='Base' LIMIT 1;
""")
ia_base = cursor.fetchone()
ia_rate = ia_base[0] if ia_base else None
print(f"\n  Taxa IA/Consumer standard/Base = {ia_rate}%")

print("\n  Simulação de taxas finais para Consumer standard/Base:")
irds_check = [
    ("IA",  "Fisico a vista (base)"),
    ("JA",  "Fisico Contactless"),
    ("HU",  "E-comm sem 3DS (RISCO ALTO)"),
    ("AU",  "E-comm com 3DS Frictionless"),
    ("AV",  "E-comm com 3DS Challenge"),
    ("AW",  "E-comm com 3DS Decoupled"),
    ("IV",  "Fisico Parcelado"),
    ("IW",  "Fisico Parcelado (alt)"),
    ("JV",  "Contactless Parcelado"),
    ("JW",  "Contactless Parcelado (alt)"),
]

for ird, desc in irds_check:
    # Taxa direta
    cursor.execute(
        "SELECT rate_pct FROM ic_base_rates WHERE ird=%s AND tier='Consumer standard' AND segment='Base' LIMIT 1;",
        (ird,)
    )
    direct = cursor.fetchone()
    if direct:
        print(f"  {ird:4s}  {desc:40s}  -> {direct[0]:.4f}% (TAXA DIRETA)")
        continue

    # Via ajuste
    cursor.execute(
        "SELECT adjustment_pct FROM ic_adjustments WHERE ird=%s AND segment='Base' LIMIT 1;",
        (ird,)
    )
    adj_row = cursor.fetchone()
    adj = adj_row[0] if adj_row else None
    if ia_rate is not None and adj is not None:
        total = ia_rate + adj
        flag = ""
        if ird in ("HU",) and adj <= 0:
            flag = "  [PROBLEMA: sem 3DS deveria ser mais caro que IA!]"
        elif ird in ("AU",) and adj > 0:
            flag = "  [PROBLEMA: com 3DS deveria ter desconto em relação ao IA!]"
        print(f"  {ird:4s}  {desc:40s}  -> {total:.4f}% (IA={ia_rate:.4f}% + adj={adj:+.4f}%){flag}")
    else:
        print(f"  {ird:4s}  {desc:40s}  -> [SEM TAXA] adj={adj}")

print("\n" + "=" * 80)
cursor.close()
conn.close()
